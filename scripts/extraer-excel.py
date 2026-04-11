"""
Extrae datos e imágenes del Excel de control de ventas de las socias.
Genera un directorio tmp-excel/ con:
  - productos.json       [{row, numero, nombre, costo, precio, vendidas}]
  - dinamicas.json       [{num, precioDefault, boletos:[{numero, nombre, telefono, pagado, precio}]}]
  - imagenes/img_{row}.{ext}   (las imágenes ancladas a cada fila de la hoja Control de Productos)
  - imagenes/manifest.json     [{row, filename}]

Uso:
  python scripts/extraer-excel.py "c:/Users/PC/Downloads/Control_Ventas_Maquillaje NAT JAZ.xlsx"
"""

import json
import sys
from pathlib import Path
import openpyxl

EXCEL_DEFAULT = r"c:/Users/PC/Downloads/Control_Ventas_Maquillaje NAT JAZ.xlsx"
OUT_DIR = Path(__file__).parent / "tmp-excel"
IMG_DIR = OUT_DIR / "imagenes"


def extract_productos(ws):
    productos = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        numero, _img, nombre, costo, precio, vendidas, *_ = row
        # Acepta filas con nombre aunque no tengan número (restocks sin numerar)
        if nombre is None:
            continue
        productos.append(
            {
                "row": row_idx,
                "numero": numero,
                "nombre": str(nombre).strip(),
                "costo": float(costo) if costo is not None else None,
                "precio": float(precio) if precio is not None else None,
                "vendidas": int(vendidas) if vendidas is not None else 0,
            }
        )
    return productos


def extract_imagenes(ws):
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    manifest = []
    for idx, img in enumerate(ws._images):
        try:
            anchor = img.anchor
            row = anchor._from.row + 1  # openpyxl es 0-indexed aquí
            data = img._data()
            ext = (img.format or "png").lower()
            filename = f"img_row{row:03d}_i{idx:03d}.{ext}"
            (IMG_DIR / filename).write_bytes(data)
            manifest.append({"row": row, "filename": filename, "bytes": len(data)})
        except Exception as e:
            print(f"  [warn] imagen #{idx}: {e}")
    # Guardar manifest
    (IMG_DIR / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return manifest


def extract_dinamica(ws, num):
    """
    Formato de hojas DINAMICA X:
      Fila 1: titulo
      Fila 4: headers (No. Rifa, Nombre, No. Telefono|PAGADO, PAGADO|PRECIO, PRECIO)
      Fila 5+: boletos
    D3 tiene formato diferente: sin columna teléfono.
    """
    boletos = []
    has_phone = True
    # Detectar si la hoja tiene columna de teléfono
    header_row = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
    if header_row[2] == "PAGADO":
        has_phone = False

    precio_default = None
    for row in ws.iter_rows(min_row=5, values_only=True):
        first = row[0]
        if not isinstance(first, int):
            continue
        if has_phone:
            numero, nombre, telefono, pagado, precio = row[:5]
        else:
            numero, nombre, pagado, precio = row[:4]
            telefono = None

        nombre_str = None
        if nombre is not None and str(nombre).strip():
            nombre_str = str(nombre).strip()
        telefono_str = None
        if telefono is not None:
            telefono_str = str(telefono).strip()
        pagado_bool = (
            str(pagado).strip().upper().startswith("PAGA")
            if pagado is not None
            else False
        )
        precio_num = float(precio) if precio is not None else None
        if precio_num is not None and precio_default is None:
            precio_default = precio_num

        boletos.append(
            {
                "numero": numero,
                "nombre": nombre_str,
                "telefono": telefono_str,
                "pagado": pagado_bool,
                "precio": precio_num,
            }
        )

    # Total de boletos: máximo número encontrado
    total = max((b["numero"] for b in boletos), default=0)
    return {
        "num": num,
        "precioDefault": precio_default,
        "totalBoletos": total,
        "boletos": boletos,
    }


def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_DEFAULT
    if not Path(excel_path).exists():
        print(f"No existe: {excel_path}")
        sys.exit(1)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Leyendo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Productos
    ws_prod = wb["Control de Productos"]
    productos = extract_productos(ws_prod)
    (OUT_DIR / "productos.json").write_text(
        json.dumps(productos, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  productos: {len(productos)} -> productos.json")

    # Imágenes de productos
    manifest = extract_imagenes(ws_prod)
    print(f"  imágenes: {len(manifest)} -> imagenes/")

    # Dinámicas
    dinamicas = []
    for i, sheet_name in enumerate(["DINAMICA 1", "DINAMICA 2", "DINAMICA 3"], start=1):
        if sheet_name not in wb.sheetnames:
            continue
        d = extract_dinamica(wb[sheet_name], i)
        dinamicas.append(d)
        llenos = sum(1 for b in d["boletos"] if b["nombre"])
        print(
            f"  dinamica {i}: total={d['totalBoletos']} "
            f"llenos={llenos} precio=${d['precioDefault']}"
        )
    (OUT_DIR / "dinamicas.json").write_text(
        json.dumps(dinamicas, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"\nListo. Output en: {OUT_DIR}")


if __name__ == "__main__":
    main()
